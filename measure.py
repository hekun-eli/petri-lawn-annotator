#!/usr/bin/env python3
"""
从 result_combined 标注图测量扩散距离。
识别红色纸片中心 → 绿色菌苔最远边界点
黄色连线必须全部在绿色区域内部
输出标注图 + Excel
"""

import os, sys, math
import cv2
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

PAPER_DIAMETER_MM = 8.0
LINE_WIDTH = 2


def line_inside(x1, y1, x2, y2, interior, green_line, steps=100):
    """
    检查线段是否在绿色区域内部。
    起点和终点之间的所有点（排除终点附近 green_line 厚度）
    都必须位于 interior 内部且不在 green_line 上。
    """
    exclude = 4  # 排除最后 4% 的终点附近（green line 厚度）
    for i, t in enumerate(np.linspace(0, 1, steps + 1)):
        x = int(x1 + (x2 - x1) * t)
        y = int(y1 + (y2 - y1) * t)
        if not (0 <= x < interior.shape[1] and 0 <= y < interior.shape[0]):
            return False
        # 终点允许在 green_line 上
        if t >= 1.0 - exclude / steps:
            continue
        # 中间点必须在 interior 内，且不在 green_line 上
        if interior[y, x] == 0 or green_line[y, x] > 0:
            return False
    return True


def main():
    default_input = (
        "/Volumes/科研 2/mcr-1 影响菌毛系统的表达/游泳运动"
        "/20250511 加药后游泳运动/result_combined"
    )
    default_output = os.path.normpath(
        os.path.join(default_input, "..", "result_measure")
    )

    in_dir = sys.argv[1] if len(sys.argv) > 1 else default_input
    out_dir = sys.argv[2] if len(sys.argv) > 2 else default_output
    in_path, out_path = Path(in_dir), Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    files = sorted([p for p in in_path.glob("*.png") if not p.stem.startswith("._")])
    print(f"输入: {in_path.resolve()}")
    print(f"输出: {out_path.resolve()}\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "扩散距离"
    headers = ["文件名", "纸片圆心X", "纸片圆心Y", "纸片半径(px)",
               "纸片直径(px)", "纸片直径(mm)", "比例(px/mm)",
               "最远点X", "最远点Y",
               "最远距离(px)", "最远距离(mm)", "最远角度(°)",
               "菌苔面积(px²)", "菌苔面积(mm²)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')

    ok = fail = 0
    for i, f in enumerate(files, 1):
        img = cv2.imread(str(f))
        if img is None:
            print(f"  [{i}/{len(files)}] ⚠️  {f.name}: 跳过")
            fail += 1; continue

        h, w = img.shape[:2]
        result = img.copy()

        # ── 检测红色纸片圆形 ──
        red = (img[:,:,2] > 200) & (img[:,:,1] < 80) & (img[:,:,0] < 80)
        if np.sum(red) < 50:
            print(f"  [{i}/{len(files)}] ❌ {f.name}: 未检测到红色纸片圆")
            fail += 1; continue

        ys, xs = np.where(red)
        pts = np.column_stack([xs, ys]).astype(np.float32)
        (pcx, pcy), pr = cv2.minEnclosingCircle(pts)
        pcx, pcy, pr = int(pcx), int(pcy), int(pr)
        px_per_mm = pr * 2 / PAPER_DIAMETER_MM

        # ── 检测绿色菌苔边界 ──
        green = (img[:,:,1] > 200) & (img[:,:,0] < 80) & (img[:,:,2] < 80)
        if np.sum(green) < 50:
            print(f"  [{i}/{len(files)}] ❌ {f.name}: 未检测到绿色菌苔边界")
            fail += 1; continue

        # 形态学连接绿色线条
        green_bin = np.uint8(green) * 255
        k = np.ones((3,3), np.uint8)
        green_bin = cv2.morphologyEx(green_bin, cv2.MORPH_CLOSE, k)

        # 找绿色轮廓（取最长的 = 菌苔边界）
        cnts, _ = cv2.findContours(green_bin, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not cnts:
            print(f"  [{i}/{len(files)}] ❌ {f.name}: 绿色轮廓提取失败")
            fail += 1; continue

        lawn = max(cnts, key=cv2.contourArea)
        lawn_area_px = int(cv2.contourArea(lawn))

        # 创建绿色内部实心蒙版（flood fill 法）
        boundary_img = np.uint8(green) * 255
        boundary_img = cv2.dilate(boundary_img, np.ones((3,3),np.uint8), iterations=1)
        interior_filled = boundary_img.copy()
        interior_mask = np.zeros((h+2, w+2), dtype=np.uint8)
        cv2.floodFill(interior_filled, interior_mask, (w//2, h//2), 255)
        interior_filled[boundary_img > 0] = 0  # 去掉绿色线条本身

        # ── 找最远点（线段全程在绿色内部，不碰绿线） ──
        max_dist = 0
        fx, fy = pcx, pcy

        # 简化轮廓点加快速度
        peri = cv2.arcLength(lawn, True)
        simplified = cv2.approxPolyDP(lawn, max(1, peri * 0.002), True)

        for pt in simplified:
            px, py = pt[0]
            d = math.hypot(px - pcx, py - pcy)
            if d > max_dist and line_inside(pcx, pcy, px, py, interior_filled, boundary_img):
                max_dist = d
                fx, fy = px, py

        if max_dist == 0:
            # 后备：放宽检查（只要求不在 green_line 上）
            for pt in simplified:
                px, py = pt[0]
                d = math.hypot(px - pcx, py - pcy)
                if d <= max_dist:
                    continue
                # 简单检查：中间点不落在 green_line 上
                ok = True
                for t in np.linspace(0, 1, 51):
                    x = int(pcx + (px-pcx)*t)
                    y = int(pcy + (py-pcy)*t)
                    if t < 0.95 and boundary_img[y,x] > 0:
                        ok = False; break
                if ok:
                    max_dist = d
                    fx, fy = px, py

        max_dist_mm = round(max_dist / px_per_mm, 2)
        angle = round(math.degrees(math.atan2(fy - pcy, fx - pcx)), 1)

        # ── 绘制（黄色线段提前 6px 终止，避免与绿线重叠） ──
        # 计算提前终止点
        angle_rad = math.atan2(fy - pcy, fx - pcx)
        gap = 15  # 提前 px 数（避免与 4px 绿色边界重叠）
        ex = fx - int(gap * math.cos(angle_rad))
        ey = fy - int(gap * math.sin(angle_rad))
        cv2.line(result, (pcx, pcy), (ex, ey), (0, 255, 255), LINE_WIDTH, cv2.LINE_AA)
        cv2.circle(result, (pcx, pcy), 4, (255, 0, 0), -1)
        cv2.circle(result, (ex, ey), 4, (0, 255, 255), -1)

        cv2.imwrite(str(out_path / f.name), result)

        # ── Excel ──
        row = [f.name, pcx, pcy, pr, pr * 2, PAPER_DIAMETER_MM,
               round(px_per_mm, 2), fx, fy,
               round(max_dist, 1), max_dist_mm, angle,
               lawn_area_px, round(lawn_area_px / (px_per_mm ** 2), 1)]
        for col, val in enumerate(row, 1):
            ws.cell(row=i + 1, column=col, value=val)

        ok += 1
        if i == 1 or i % 10 == 0 or i == len(files):
            print(f"  [{i}/{len(files)}] ✅ {f.name}: "
                  f"纸片r={pr}px Ø={pr*2}px "
                  f"最远={round(max_dist,1)}px({max_dist_mm}mm) ∠{angle}°")

    # ── Excel 保存 ──
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
    wb.save(str(out_path / "measure_results.xlsx"))

    print(f"\n📊 Excel: {out_path / 'measure_results.xlsx'}")
    print(f"🎯 完成: 成功 {ok} 张，失败 {fail} 张")


if __name__ == "__main__":
    main()
